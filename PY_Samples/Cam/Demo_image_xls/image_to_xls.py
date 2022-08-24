#-------------------------------------------------------------------------------
# Name:        module1
# Purpose:
#
# Author:      SESA237770
#
# Created:     30.07.2022
# Copyright:   (c) SESA237770 2022
# Licence:     <your licence>
#-------------------------------------------------------------------------------

import openpyxl
from openpyxl.styles import PatternFill
from PIL import Image
from numpy import asarray


#==================================================================
def rgb_to_hex(rgb):
        return '%02x%02x%02x' % rgb

def hex_to_rgb(value):
        value = value.lstrip('#')
        lv = len(value)
        return tuple(int(value[i:i+lv//3], 16) for i in range(0, lv, lv//3))
#==================================================================

def main():


#==================================================================
    image = Image.open('image.png')
    imageXLSX="image.xlsx"
    # summarize some details about the image
    print(image.format)
    print(image.size)
    print(image.mode)




    # asarray() class is used to convert
    # PIL images into NumPy arrays
    numpydata = asarray(image)

    # <class 'numpy.ndarray'>
    print(type(numpydata))

    #  shape
    print(numpydata.shape)




#==================================================================
#==================================================================
#==================================================================


#==================================================================
#==================================================================
#==================================================================
    wb = openpyxl.load_workbook(imageXLSX)
    ws = wb['image'] #Name of the working sheet

    fill_cell1 = PatternFill(patternType='solid',fgColor='FC2C03')
    fill_cell2 = PatternFill(patternType='solid',fgColor='03FCF4')
    fill_cell3 = PatternFill(patternType='solid',fgColor='35FC03')
    fill_cell4 = PatternFill(patternType='solid',fgColor='FCBA03')

    ws.cell(row=1, column=1).fill = fill_cell1
    ws['B3'].fill = fill_cell2
    ws['B4'].fill = fill_cell3
    ws['B5'].fill = fill_cell4

#==================================================================
#==================================================================
#==================================================================
    print('===BEGIN COLOR LOOP===')
        # data
##    print(numpydata)
    nrow=0
    ncol=0
    for X in numpydata:
        nrow+=1
        ncol=0
        for Y in X:
            ncol+=1


            sRGB='FFFFFF'
            sRGB=rgb_to_hex((Y[0], Y[1], Y[2]))
            sRGB=sRGB.upper()

##            print(nrow,ncol,'sRGB=',sRGB,'R G B =',Y[0], Y[1], Y[2])

##            ws.cell(row=nrow, column=ncol).value = Y
            ws.cell(row=nrow, column=ncol).value = sRGB

            #===works with image size = 300 x 150
            del fill_cell1
            fill_cell1 = PatternFill(patternType='solid',fgColor=sRGB)
            ws.cell(row=nrow, column=ncol).fill = fill_cell1


            if 1==2 :
##                fill_cell1 = PatternFill(patternType='solid',fgColor=sRGB)
##                ws.cell(row=nrow, column=ncol).fill = PatternFill("solid",sRGB)
##                sheet['A1'].fill = PatternFill(bgColor=sRGB, fill_type = "solid")
                ws.cell(row=nrow, column=ncol).fill = PatternFill(start_color=sRGB,end_color=sRGB, fill_type = 'solid')





##            print('---------')

#==================================================================
#==================================================================
    wb.save(imageXLSX)

    pass

#==================================================================
#==================================================================
if __name__ == '__main__':

    print('========================================================')
    print('IMAGE TO XLS')
    print('========================================================')

    x=hex_to_rgb('FF65BA')
    y=rgb_to_hex((255, 255, 195))

    print('rgb=',x)
    print('HEX=',y.upper())

    main()

    print('========================================================')
    print('COMLETED RUN')
    print('========================================================')