#-------------------------------------------------------------------------------
# Name:        module1
# Purpose:
#
# Author:      SESA237770
#
# Created:     03.04.2017
# Copyright:   (c) SESA237770 2017
# Licence:     <your licence>
#-------------------------------------------------------------------------------

def main():
    pass

if __name__ == '__main__':
    main()

#-------------------------------------------------------------------------------
#-------------------------------------------------------------------------------
import csv, sqlite3, os , sys , re
import codecs
import datetime
# CAX User Modules
sys.path.append('D:\\MIDDLEWARE_PYTHON\\cax_modules')
from CAX_Module_Files import cax_files
from CAX_Module_Cad import cax_cad
from CAX_Module_XLSX import cax_xlsx
##from CAX_Modules_2021 import caxlib
import HTML
import string


#@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@
#***************************************************************************

print('=====================================================')
print('-------------Tool_01_CreateTables.py-----Path Root-----------')
print('=====================================================')
nCommandParameters= len(sys.argv)
print(len(sys.argv))

print('=====ARGUMENTS========================================')
print( 'Arguments:', len(sys.argv), 'arguments.')
print('Argument:', str(sys.argv[0]))
if len(sys.argv) == 3:
    print('Argument:', str(sys.argv[1]))
    print('Argument:', str(sys.argv[2]))
print('=====================================================')

###############################################################################
print('=====DATES TIMERS====================================')
print( 'Arguments:', len(sys.argv), 'arguments.')
dt_date = datetime.datetime.now()
sToday= dt_date.strftime('%y/%m/%d %H:%M' )
sTodayTime = 'Time: ' + sToday
print(sTodayTime)

###############################################################################
### used for debuging when no command line available
dir_Folder_Export= 'MLS0011_MLS_Aenderung-bwe-trafo-BG-SESA565070'

dir_Basefolder='D:\\SYMPHONY_Download_MLS_PART\\PDM_EXPORT_02_MLS\\'
dir_Templatefolder='D:\\SYMPHONY_Download_MLS_PART\\WEBSITE-Templates'

if len(sys.argv) == 3:
    dir_path=dir_Basefolder + sys.argv[1] +  '\\SAP_Feller'
    dir_Folder_Export= sys.argv[1]
    sLEVELS=sys.argv[2]
else:
    dir_path=dir_Basefolder + dir_Folder_Export  + '\\SAP_Feller'
    sLEVELS='4'

BOM_TYPE='PART'
###############################################################################
print('Export Dir=[' + dir_path+ ']')
dir_path_root=dir_Basefolder + dir_Folder_Export
dir_path_pdf=dir_path_root   + '\\referenceFiles'

print('====Redirect STD Output===============================')
# redirect StdOut output to log file
old_stdout = sys.stdout
'D:\SYMPHONY_Download_MLS_PART\PDM_EXPORT_02_MLS\MLS0006_B0001-FUN37_VA\SAP_Feller\log'
sLogFile=dir_path_root +'\\SAP_Feller\\log\\Tool_03_ML_STRUCTURE.log'
print ('Log File=' + dir_path_root)
log_file = open(sLogFile,'w')
print ('Log File! open')
#----------------------------
if len(sys.argv)==3000:
   print ('ALL LOGS SENT TO FILE !!!!!!!!!!!!!!' )
   sys.stdout = log_file
#----------------------------
print('=====================================================')

print ('DIR PATH=' + dir_path )
print ('DIR PATH_ROOT=' + dir_path_root )

print('=====================================================')
print('=====================================================')

# OUTPUT will be created
sPathFileDB=dir_path_root + '\\SAP_Feller\\csv\\matchingnumbers_ML.db'

sPathFile_GUI_Table_Filter_HTML = dir_Templatefolder + '\\GUI_Table_Filter.html'
sPathFile_GUI_Table_Filter_SAP_UPLOAD_HTML = dir_Templatefolder + '\\GUI_Table_Filter_SAP_UPLOAD.html'

sPathFile_BOM_MATRIX_TEMPLATE = dir_Templatefolder + '\\REPORTS_001_BOM_MATRIX.html'
sPathFile_BOM_MATRIX = dir_path_root + '\\REPORTS_001_BOM_MATRIX.html'

sPathFile_BOM_MATRIX_VECTOR_VER = dir_path_root + '\\BOM_MATRIX_VECTOR_VER.html'
sPathFile_BOM_MATRIX_VECTOR_HOR = dir_path_root + '\\BOM_MATRIX_VECTOR_HOR.html'
sPathFile_BOM_MATRIX_PARENT_CHILD = dir_path_root + '\\BOM_MATRIX_PARENT_CHILD.html'
sPathFile_BOM_MATRIX_FIND_LISTS = dir_path_root + '\\BOM_MATRIX_FIND_LISTS.html'

MATRIX_VECTORS=[]
MATRIX_PARENT_CHILD=[]

#==============================================================================
# All Required DB Tables are Created Her
#==============================================================================
# con = sqlite3.connect(":memory:")
con = sqlite3.connect(sPathFileDB)
con.text_factory = str
cur = con.cursor()
print('-----SQL_MATRIX_VECTORS-------------------------------------')

cur.execute("SELECT TYPE FROM BOM_ML")
rows = cur.fetchall()

if ('COMMERCIAL REFERENCE',) in rows:
    BOM_TYPE='COMMERCIAL REFERENCE'
    print('BOM TYPE: ', BOM_TYPE)

if BOM_TYPE=='COMMERCIAL REFERENCE':

    SQL_MATRIX_HORIZONTAL='''
select
 NUMBER , MATKRTZTEXT, LEVEL
from BOM_MATRIX_VECTORS
WHERE LEVEL not like '0'
order by LEVEL ASc , NUMBER ASc
'''
    SQL_MATRIX_VERTICAL='''
select
 NUMBER , MATKRTZTEXT, LEVEL
from BOM_MATRIX_VECTORS
WHERE LEVEL not like '0'
order by LEVEL ASc
'''
##  NOTEE : LEVEL not like (SELECT MAX(CAST(LEVEL AS INTEGER))
## Levels depth limited during CSV upload


else:
    BOM_TYPE='PART'
    SQL_MATRIX_HORIZONTAL='''
select
 NUMBER , MATKRTZTEXT, LEVEL
from BOM_MATRIX_VECTORS
order by LEVEL ASc , NUMBER ASc
'''
    SQL_MATRIX_VERTICAL='''
select
 NUMBER , MATKRTZTEXT, LEVEL
from BOM_MATRIX_VECTORS
order by LEVEL ASc
'''




SQL_MATRIX_FIND_LISTS='''
select
NUMBER, FIND_LISTS
from BOM_MATRIX_FIND_NUMBERS
order by NUMBER desc
'''

cur.execute(SQL_MATRIX_HORIZONTAL)
print( SQL_MATRIX_HORIZONTAL)
print('Query result rows=',cur.rowcount)
result_Vector_Hor=cur.fetchall()

cur.execute(SQL_MATRIX_VERTICAL)
print( SQL_MATRIX_VERTICAL)
print('Query result rows=',cur.rowcount)
result_Vector_Ver=cur.fetchall()

cur.execute(SQL_MATRIX_FIND_LISTS)
print( SQL_MATRIX_FIND_LISTS)
print('Query result rows=',cur.rowcount)
result_Vector_Find=cur.fetchall()

print('-----SQL_PARENT_CHILD-------------------------------------')
SQL_PARENT_CHILD='''
select
PARENT_CHILD,LEVEL, PARENT, NUMBER, MATKRTZTEXT,MENGE, EINHEIT, SUCHBGR, REQUIRED,PARENT,TYPE
from PARENT_CHILD
'''
cur.execute(SQL_PARENT_CHILD)
print( SQL_PARENT_CHILD)
print('Query result rows=',cur.rowcount)
result_PARENT_CHILD=cur.fetchall()

con.commit()
con.close()
##---------------------------------------------------------------------------------------------
# DB CLOSED
##---------------------------------------------------------------------------------------------



print('')
print('-------REPORT MLS--------Create ??-----------------')

##########################################################################################################################################################
##  DATA PREPERATION
##########################################################################################################################################################
#==============================================================================
# REPORT  VECTOR
#==============================================================================
print('')
print('-------REPORT VECTOR VERTICAL --------Create HTML Content + Table-----------------')
HTML_VECTORS_VER=''
HTML_VECTORS_VER=HTML.table(result_Vector_Ver,header_row=['NUMBER' , 'MATKRTZTEXT', 'LEVEL'])
HTML_VECTORS_VER=HTML_VECTORS_VER.replace('<TABLE','<TABLE id="data1"')  # ID for F Java script
ExportFolder=  '<B>' + dir_Folder_Export +'</B>'
print('')
print('-------REPORT VECTOR HORIZONTAL--------Create HTML Content + Table-----------------')
HTML_VECTORS_HOR=''
HTML_VECTORS_HOR=HTML.table(result_Vector_Hor,header_row=['LEVEL', 'MATKRTZTEXT', 'NUMBER'])
HTML_VECTORS_HOR=HTML_VECTORS_HOR.replace('<TABLE','<TABLE id="data2"')  # ID for F Java script
ExportFolder=  '<B>' + dir_Folder_Export +'</B>'
print('')
print('')
print('-------REPORT FIND LISTS--------Create HTML Content + Table-----------------')
HTML_FIND_LISTS=''
HTML_FIND_LISTS=HTML.table(result_Vector_Find,header_row=['NUMBER', 'FIND_LISTS'])
HTML_FIND_LISTS=HTML_FIND_LISTS.replace('<TABLE','<TABLE id="data3"')  # ID for F Java script
ExportFolder=  '<B>' + dir_Folder_Export +'</B>'
print('')

#==============================================================================
# REPORT  PARENT_CHILD
#==============================================================================
print('')
print('-------REPORT PARENT_CHILD--------Create HTML Content + Table-----------------')
HTML_PARENT_CHILD=''
HTML_PARENT_CHILD=HTML.table(result_PARENT_CHILD,header_row=['PARENT_CHILD','LEVEL', 'PARENT', 'NUMBER',' MATKRTZTEXT','MENGE', 'EINHEIT', 'SUCHBGR', 'REQUIRED','TYPE'])
HTML_PARENT_CHILD=HTML_PARENT_CHILD.replace('<TABLE','<TABLE id="data4"')  # ID for F Java script
ExportFolder=  '<B>' + dir_Folder_Export +'</B>'
print('')

##########################################################################################################################################################
##  DATA PUBLICATION
##########################################################################################################################################################
#==============================================================================
# WRITING TO TEMPLATE HTML REPORTs
#==============================================================================

##LIST_BOM_MATRIX_Txt=HTML_VECTORS + '<BR>' + HTML_PARENT_CHILD
##sHTML_BOM_MATRIX=cax_files.read_file_to_txt(sPathFile_BOM_MATRIX_TEMPLATE)
##sHTML_BOM_MATRIX=sHTML_BOM_MATRIX.replace('#REPORT_TABLE#',LIST_BOM_MATRIX_Txt)
##sHTML_BOM_MATRIX=sHTML_BOM_MATRIX.replace('#TIME#',sToday)
##sHTML_BOM_MATRIX=sHTML_BOM_MATRIX.replace('#EXPORT#',dir_Folder_Export)
##cax_files.write_txt_to_file(sPathFile_BOM_MATRIX,sHTML_BOM_MATRIX)

cax_files.write_txt_to_file(sPathFile_BOM_MATRIX_PARENT_CHILD,HTML_PARENT_CHILD )
cax_files.write_txt_to_file(sPathFile_BOM_MATRIX_VECTOR_VER,HTML_VECTORS_VER)
cax_files.write_txt_to_file(sPathFile_BOM_MATRIX_VECTOR_HOR,HTML_VECTORS_HOR)
cax_files.write_txt_to_file(sPathFile_BOM_MATRIX_FIND_LISTS,HTML_FIND_LISTS)

sPathFile_BOM_MATRIX_PARENT_CHILD

#==============================================================================
# MAKE EXCEL
#==============================================================================

##---------------------------------------------------------------------------------------------
# Make XLS
##---------------------------------------------------------------------------------------------
from shutil import copyfile


sPathFileXLS_TEMPLATE='C:\Schneider\ScriptingPath\PLUGINS_BOM_MATRIX\TEMPLATES\BOM_MATRIX_MLS_PYTHON_TEMPLATE.xlsx'
print(sPathFileXLS_TEMPLATE)

sPathFileXLS=dir_path_root + '\\SAP_Feller\\BomMatrix\\BM_'+ dir_Folder_Export + '_RAW.XLSX'
sPathFileXLS_USER=dir_path_root + '\\SAP_Feller\\BomMatrix\\BM_'+ dir_Folder_Export + '.XLSX'

# sPathFileXLS=dir_path_root + '\\SAP_Feller\\BomMatrix\\BOM_MATRIX_MLS_PYTHON.XLSX'

print(sPathFileXLS)

# copyfile(src, dst)
copyfile(sPathFileXLS_TEMPLATE, sPathFileXLS)

print('===============================================================')
# RNG_Artikel_Description_Sheet1
# RNG_ARTIKEL_Sheet1
# RNG_Bezeichnung_TXT_Sheet1
# RNG_BG_PDM_Sheet1
# RNG_BG_Sheet1
# RNG_COLOR_Sheet1
# RNG_IMAGES_Sheet1
# RNG_LEVEL_H_Sheet1
# RNG_LEVEL_V_Sheet1
# RNG_LINK_ARTIKEL_MATERIAL_Sheet1
# RNG_MATNR_Bricks_Sheet1
# RNG_PD_COLOR_Sheet1
# RNG_PD_PDM_Sheet1
# RNG_PD_Sheet1
# RNG_PDM_STATUS_Sheet1
# RNG_SAP_PDM_USER_Sheet1
# RNG_SAP_STATUS_Sheet1
# RNG_SUCHBEGRIFF_Sheet1
print('====OPEN WORKBOOK XLSX')
WB_Template = cax_xlsx.OpenWorkbook(sPathFileXLS_TEMPLATE)
print('=====populating XLS================================================')
lList=['START','s','d','f','g','h','j','s','d','f','g','h','LAST']

Sheet_BOM = WB_Template.get_sheet_by_name('BOM_MATIX_PYTHON')
Sheet_PC = WB_Template.get_sheet_by_name('PARENT-CHILD')
Sheet_FIND = WB_Template.get_sheet_by_name('FINDLISTS')

Sheet_BOM['C1'] = dir_Folder_Export

# !!!!!!!!!!!!!!!!!!!!!Sheet corrupted with the formula auto repair in xls is starte
# Find Number
##Sheet_BOM['P13'] = '''=IFERROR(VLOOKUP(P$4;'FINDLISTS'!$A:$B;2;FALSE);"'")'''
# PARENT-CHILD
##Sheet_BOM['P18'] = '''=IFERROR(VLOOKUP($E18&">"&P$4;'PARENT-CHILD'!$A:$I;6;FALSE);"'")'''


print('=====populating XLS with PARENT CHILD LOOKUP===========================')
# result_PARENT_CHILD

print(result_PARENT_CHILD)
nRowsPC = len(result_PARENT_CHILD)
nItemsPC =len(result_PARENT_CHILD[1])
print('ROWS:  ' , nRowsPC)
print('ITEMS:  ' , nItemsPC)

for i in range(2,nRowsPC+2):     ## row 1 has heading in XLS
    print('ROWS:  ' , i, result_PARENT_CHILD[i-2][0])
    for j in range(0,nItemsPC):
        cellref=Sheet_PC.cell(row=i, column=1+j)
##        cellref.value=str(i) +'  ' + str(j)
        cellref.value=result_PARENT_CHILD[i-2][j]

print('=====populating XLS with FIND LISTS LOOKUP=============================')
# result_PARENT_CHILD

print(result_Vector_Find)
nRowsF = len(result_Vector_Find)
nItemsF =len(result_Vector_Find[1])
print('ROWS:  ' , nRowsF)
print('ITEMS:  ' , nItemsF)

for i in range(2,nRowsF+2):      ## row 1 has heading
    for j in range(0,nItemsF):
        cellref=Sheet_FIND.cell(row=i, column=1+j)
##        cellref.value=str(i) +'  ' + str(j)
        cellref.value=result_Vector_Find[i-2][j]
##        print('value:  ' , result_Vector_Find[i-2][j])


print('======result_Vector_Hor====NUMBER MATKRTZTEXT LEVEL=============')
lList = [lList[0] for lList in result_Vector_Hor]
nRangeRow=0   # 0, 1 , ......
sRangeName='RNG_SAP_PDM_USER_Sheet1'
WB_New=cax_xlsx.FillRange(nRangeRow,sRangeName,WB_Template,lList)

lList = [lList[2] for lList in result_Vector_Hor]
nRangeRow=0   # 0, 1 , ......
sRangeName='RNG_LEVEL_H_Sheet1'
WB_New=cax_xlsx.FillRange(nRangeRow,sRangeName,WB_Template,lList)

lList = [lList[1] for lList in result_Vector_Hor]
nRangeRow=0   # 0, 1 , ......
sRangeName='RNG_Bezeichnung_TXT_Sheet1'
WB_New=cax_xlsx.FillRange(nRangeRow,sRangeName,WB_Template,lList)



print('=======result_Vector_Ver======NUMBER MATKRTZTEXT LEVEL=========')
lList = [lList[0] for lList in result_Vector_Ver]
nRangeRow=0   # 0, 1 , ......
sRangeName='RNG_ARTIKEL_Sheet1'
sRangeName='RNG_PARENT_Sheet1'
sRangeName='RNG_BG_PDM_Sheet1'


WB_New=cax_xlsx.FillRange(nRangeRow,sRangeName,WB_New,lList)

lList = [lList[2] for lList in result_Vector_Ver]
nRangeRow=0   # 0, 1 , ......
sRangeName='RNG_LEVEL_V_Sheet1'
WB_New=cax_xlsx.FillRange(nRangeRow,sRangeName,WB_New,lList)

lList = [lList[1] for lList in result_Vector_Ver]
nRangeRow=0   # 0, 1 , ......
sRangeName='RNG_Artikel_Description_Sheet1'
WB_New=cax_xlsx.FillRange(nRangeRow,sRangeName,WB_New,lList)


print('=======FIND NUBER======NUMBER FIND_LISTS=========')
##lList = [lList[1] for lList in result_Vector_Find]
##nRangeRow=0   # 0, 1 , ......
##sRangeName='RNG_SUCHBEGRIFF_Sheet1'
##WB_New=cax_xlsx.FillRange(nRangeRow,sRangeName,WB_New,lList)

print('=======Insert colums and rows as LEVEL seperators=========')
## colum 15 row 2 inset columns
## row 17 column 2
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
print('=======Find Level Transitions COLUMNS=========line 416') #   'P = col 16'
lListInsertedCol=[]
Level_Horizontal=""
# 2022 03 07   nCol_Start=16 + 0
nCol_Start=20 + 0
nCol=nCol_Start
##LEVEL=Sheet_BOM.cell(row=2, column=17).value
LEVEL=Sheet_BOM.cell(row=2, column=21).value
LEVEL_PRE=''
while (LEVEL !=None):
    nCol +=1
    LEVEL=Sheet_BOM.cell(row=2, column=nCol).value
    if LEVEL != LEVEL_PRE :  ##  and LEVEL!='x1' x to disable
##        Sheet_BOM.insert_cols(nCol) ## insert before
        lListInsertedCol.append(nCol)
##        nCol +=1
        dummy = 1
    LEVEL_PRE=LEVEL
    print('Column:  ',LEVEL)

print('=======Find Level Transitions ROWS=========') # 'A17'
lListInsertedRow=[]
Level_Verical=""
nRow_Start=17 + 0
nRow=nRow_Start
LEVEL=Sheet_BOM.cell(row=nRow, column=2).value
LEVEL_PRE=''
while (LEVEL !=None):
    nRow +=1
    LEVEL=Sheet_BOM.cell(row=nRow, column=2).value
    if LEVEL != LEVEL_PRE :   ## and LEVEL!='x0' x to disable
##        Sheet_BOM.insert_rows(nRow) ## insert before
         lListInsertedRow.append(nRow) ## insert before
##        nRow +=1
         Dummy=1
    LEVEL_PRE=LEVEL
    print('ROW: ',LEVEL)


print('=======COLOR Edit Inserted Columns=========')
##https://www.pythonexcel.com/openpyxl-write-to-cell.php
from openpyxl.styles import Alignment

print(lListInsertedCol)
nLevel=0
for nCol in lListInsertedCol:
        nLevel+=1
        for x in range(1, max(lListInsertedCol)):
            Sheet_BOM.cell(row=x, column=nCol).fill = PatternFill("solid", fgColor="3366FF") #  https://openpyxl.readthedocs.io/en/stable/styles.html
        cellref1=Sheet_BOM.cell(row=3, column=nCol)
##        cellref1.value="LEVEL: " + str(nLevel)
        cellref1.alignment = Alignment(textRotation=90)
        cellref1.alignment = Alignment(horizontal='center' , vertical='center', textRotation=90)
##Sheet_BOM.delete_cols(nCol +1  , 242+ nCol_Start -nCol)  ## 227 Rows,  242 Cols , Rows until END
## seting properties of cells in not acumalative, properites must be

print('=======Set Width Inserted Columns=========')
##for nCol in lListInsertedCol:
##        sColumnLetter =get_column_letter(nCol)
##        print(sColumnLetter)
##        Sheet_BOM.column_dimensions[get_column_letter(nCol)].width = 5
##        Sheet_BOM.column_dimensions['AN'].width = 5


print('=======COLOR Edit Inserted  Rows=========')
print(lListInsertedRow)
nLevel=-1
for nRow in lListInsertedRow:
    nLevel+=1
##    for x in range(1, 16):
    for x in range(1, max(lListInsertedRow)+3):
        #  https://openpyxl.readthedocs.io/en/stable/styles.html
                Sheet_BOM.cell(row=nRow, column=x + 1).fill = PatternFill("solid", fgColor="3366FF") # x + 3 coler a few columns to little ? matrix not perfect square
    cellref=Sheet_BOM.cell(row=nRow, column=3)
##    cellref.value="LEVEL: " + str(nLevel)

##Sheet_BOM.delete_rows(nRow +1  , nRow_Start + 227-nRow)  ## 227 Rows,  242 Cols , Rows until END

print('=======Group Columns in EXCEL Colapsing Columns=========line 494')
import zipapp
Sheet_BOM.sheet_properties.outlinePr.summaryRight = False
##Sheet_BOM.column_dimensions.group(get_column_letter(25),get_column_letter(30))
Group_List=zip(lListInsertedCol,lListInsertedCol[1:] )
##print(Group_List)
Group_Col =list(Group_List)
print(Group_Col)
## if  [2, 4, 6, 8]  [1, 3, 5, 7, 9] then [(1, 2), (3, 4), (5, 6), (7, 8)]
print('Grouping.....')
for Group in Group_Col:
    nStartCoumn=Group [0] + 1
    nEndColumn=Group [1] -1
    print(nStartCoumn , nEndColumn)
    if nEndColumn -nStartCoumn > 2:  # Min GROUP Size
        print('Group: ' ,get_column_letter(nStartCoumn ), get_column_letter(nEndColumn))
        Sheet_BOM.column_dimensions.group(get_column_letter(nStartCoumn),get_column_letter(nEndColumn))

print('=======Group Rows n EXCEL Colapsing Rows=========')
Group_List=zip(lListInsertedRow,lListInsertedRow[1:] )
##print(Group_List)
Group_Row =list(Group_List)
print(Group_Row)
## if  [2, 4, 6, 8]  [1, 3, 5, 7, 9] then [(1, 2), (3, 4), (5, 6), (7, 8)]
print('Grouping.....')
for Group in Group_Row:
    nStartRow=Group [0] + 1
    nEndRow=Group [1] -1
    print(nStartRow , nEndRow)
    if nEndRow -nStartRow > 2:   # Min GROUP Size
        print('Group Row: ' ,nStartRow , nEndRow)
        Sheet_BOM.row_dimensions.group(nStartRow,nEndRow)
##=======================================================================================
## 2022 03 07 inserted 4 Columns P will now be T
## 2022 03 07 inserted 4 Columns R (PQRS) will now be (TUVW) V
## 2022 03 07 inserted 4 Columns S (PQRS)will now be (TUVW) W
## Unchanged Columns  F G
##  P16 = T16
##  P17 = T17
##  P16 = T16


print('=======Add formula=========')
##=+WENNFEHLER(SVERWEIS($J18&">"&Q$4;'PARENT-CHILD'!$A:$I;6;FALSCH);"'")
from openpyxl.formula.translate import Translator
sFormula =Sheet_BOM['T17'].value
print(sFormula)
sFormula2 =Sheet_BOM['T16'].value
print(sFormula2)
## cell = Q18
##sFormula ='=+IFERROR(VLOOKUP($J18&">"&Q$4,$\'PARENT-CHILD\'.$A:$I,6,FALSE()),"\'")'
##print(sFormula)
# move the formula one colum to the right
Sheet_BOM['V13'] = Translator(sFormula2, "T16").translate_formula("V13")
Sheet_BOM['W13'] = Translator(sFormula2, "T16").translate_formula("W13")

Sheet_BOM['V19'] = Translator(sFormula, "T17").translate_formula("V19")
Sheet_BOM['W19'] = Translator(sFormula, "T17").translate_formula("W19")

print('=============================================================')
print('=======Add formula to TopLeft Parent-Parent of Matrix=========')
print('=============================================================')
Group_2D_List_temp=zip(lListInsertedRow,lListInsertedCol )
Group_2D_List_Parent_Parent =list(Group_2D_List_temp)
print(Group_2D_List_Parent_Parent)

sFormula =Sheet_BOM['T17'].value
sFormula_FinNumber =Sheet_BOM['T16'].value
sFormula_CR =Sheet_BOM['F17'].value
sFormula_PD =Sheet_BOM['G17'].value

Group_2D_List_Parent_Parent_modified = Group_2D_List_Parent_Parent[1:]
Group_2D_List_Parent_Parent_modified = Group_2D_List_Parent_Parent[0:-1]  ## include first and last element
print('------------------------------------------------')
print('Matrix Groups: ',Group_2D_List_Parent_Parent)
print('Matrix Groups: ',Group_2D_List_Parent_Parent_modified)

for Group_2D in Group_2D_List_Parent_Parent_modified:
    nRow=Group_2D [0] + 1
    nCol=Group_2D [1] + 1
    print('')
    print(nRow , nCol)

## Write FORMULA to BEGIN of Sub Matrix (levels N to N+1)  Parent child= Level n to Level n
    ##  First Formula
    sCellEnd= get_column_letter(nCol) + str(nRow)
    Sheet_BOM.cell(row=nRow, column=nCol).value=sCellEnd   #DEBUG
    Sheet_BOM[sCellEnd] = Translator(sFormula, "T17").translate_formula(sCellEnd)
    ##  Second Formula
    sCellEnd= get_column_letter(nCol+1) + str(nRow)
    Sheet_BOM.cell(row=nRow, column=nCol+1).value=sCellEnd   #DEBUG
    Sheet_BOM[sCellEnd] = Translator(sFormula, "T17").translate_formula(sCellEnd)

## Find Number Row =13formula in P16
    ##  First Formula
    sCellEnd= get_column_letter(nCol+0) + str(13)
    Sheet_BOM[sCellEnd] = Translator(sFormula_FinNumber, "T16").translate_formula(sCellEnd)
    ##  Second Formula
    sCellEnd= get_column_letter(nCol+1) + str(13)
    Sheet_BOM[sCellEnd] = Translator(sFormula_FinNumber, "T16").translate_formula(sCellEnd)

#### PD Lookup formulaRow =17 formula in F17
##    ##  First Formula  PD Lookup
##    sCellEnd= get_column_letter(6) + str(nRow)
##    Sheet_BOM[sCellEnd] = Translator(sFormula_CR, "F17").translate_formula(sCellEnd)
##    ##  Second Formula  BG Lookup
##    sCellEnd= get_column_letter(7) + str(nRow)
##    Sheet_BOM[sCellEnd] = Translator(sFormula_PD, "G17").translate_formula(sCellEnd)

## Copy cell formating to  BEGIN of Sub Matrix
    sCellSourceFormat1= get_column_letter(nCol) + str(5)  # contains formating
    sCellSourceFormat2= get_column_letter(nCol+1) + str(5)  # contains formating
    sCellDestinationFormat1= get_column_letter(nCol) + str(nRow)
    sCellDestinationFormat2= get_column_letter(nCol+1) + str(nRow)
    c1=Sheet_BOM[sCellSourceFormat1]
    c2=Sheet_BOM[sCellSourceFormat2]
    Sheet_BOM[sCellDestinationFormat1]._style=c1._style
    Sheet_BOM[sCellDestinationFormat2]._style=c2._style
    print('Format from: ',sCellSourceFormat1,sCellDestinationFormat1)


## Write Formula to  BEGIN of Parent Vector
    sCellCR= 'F' + str(nRow)
    sFormulaCR =Sheet_BOM['F17'].value

    sCellPD= 'G' + str(nRow)
    sFormulaPD =Sheet_BOM['G17'].value

    sCellPD_SAP= 'H' + str(nRow)
    sFormulaPD_SAP =Sheet_BOM['H17'].value

    sCellBG= 'J' + str(nRow)
##  sCellBG is constant from BOM MATRIX

    sCellBG_SAP= 'K' + str(nRow)
    sFormulaBG_SAP =Sheet_BOM['K17'].value

    print('CR PD Cells: ',sCellCR , sCellPD, sCellBG, sCellPD_SAP,  sCellBG_SAP)
    print(sFormulaCR)
    print(sFormulaPD)
    print(sFormulaPD_SAP)
    print(sFormulaBG_SAP)

    Sheet_BOM[sCellCR] = Translator(sFormulaCR, "F17").translate_formula(sCellCR)
    Sheet_BOM[sCellPD] = Translator(sFormulaPD, "G17").translate_formula(sCellPD)
    Sheet_BOM[sCellPD_SAP] = Translator(sFormulaPD_SAP, "H17").translate_formula(sCellPD_SAP)
    Sheet_BOM[sCellBG_SAP] = Translator(sFormulaBG_SAP, "K17").translate_formula(sCellBG_SAP)

print('=============================================================')
print('=======Add formula to TopLeft PARENT-CHILD of Matrix=========')
print('=============================================================')
Group_2D_List_temp=zip(lListInsertedRow,lListInsertedCol[1:] )
Group_2D_List_Parent_Child =list(Group_2D_List_temp)
print(Group_2D_List_Parent_Child)
Group_2D_List_Parent_Child_modified = Group_2D_List_Parent_Child[0:-1]  ## include first and last element
print('------------------------------------------------')
print('Matrix Groups: ',Group_2D_List_Parent_Child)
print('Matrix Groups: ',Group_2D_List_Parent_Child_modified)

for Group_2D in Group_2D_List_Parent_Child_modified:
    nRow=Group_2D [0] + 1
    nCol=Group_2D [1] + 1
    print('')
    print(nRow , nCol)

## Write FORMULA to BEGIN of Parent_Child Sub Matrix (levels N to N+1)  Parent child= Level n to Level n
    ##  First Formula
    sCellEnd= get_column_letter(nCol) + str(nRow)
    Sheet_BOM.cell(row=nRow, column=nCol).value=sCellEnd   #DEBUG
    Sheet_BOM[sCellEnd] = Translator(sFormula, "T17").translate_formula(sCellEnd)
    ##  Second Formula
    sCellEnd= get_column_letter(nCol+1) + str(nRow)
    Sheet_BOM.cell(row=nRow, column=nCol+1).value=sCellEnd   #DEBUG
    Sheet_BOM[sCellEnd] = Translator(sFormula, "T17").translate_formula(sCellEnd)

    ## Copy cell formating to  BEGIN of Parent_Child Sub Matrix
    sCellSourceFormat1= get_column_letter(nCol) + str(4)  # contains formating
    sCellSourceFormat2= get_column_letter(nCol+1) + str(4)  # contains formating
    sCellDestinationFormat1= get_column_letter(nCol) + str(nRow)
    sCellDestinationFormat2= get_column_letter(nCol+1) + str(nRow)
    c1=Sheet_BOM[sCellSourceFormat1]
    c2=Sheet_BOM[sCellSourceFormat2]
    Sheet_BOM[sCellDestinationFormat1]._style=c1._style
    Sheet_BOM[sCellDestinationFormat2]._style=c2._style
    print('Format from: ',sCellSourceFormat1,sCellDestinationFormat1)


print('=============================================================')
print('=============================================================')
print('=======Populate FROM TopLeft TO BottomRight of Matrix=========')
Martices_List=zip(lListInsertedRow,lListInsertedCol )
Martices_List =list(Martices_List)
print(Martices_List)

print('===============================================================')
print('====SAVE WORKBOOK XLSX')
cax_xlsx.SaveWorkbook(sPathFileXLS,WB_New)

from os.path import exists
file_exists = exists(sPathFileXLS_USER)
if not file_exists :
    cax_xlsx.SaveWorkbook(sPathFileXLS_USER,WB_New)


print('===============================================================')
print('====EXECUTION COMPLETE===')
print('===============================================================')

print('---------------------------------------------')
print('-------Run Complete-------------------------')
print('---------------------------------------------')
print('---Summary---------------------------------------')
if len(sys.argv) == 3:
    print('EXPORT:', str(sys.argv[1]))
    print('LEVELS:', str(sys.argv[2]))
    print('BOM TYPE: ', BOM_TYPE)
print('---------------------------------------------')

# Close Log file
#sys.stdout = old_stdout
log_file.close()






